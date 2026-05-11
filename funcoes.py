import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def salvar_dados_excel(arquivo_upload, dados_novos, nome_aba):
    # 1. Preparar os dados novos
    df_novo = pd.DataFrame(dados_novos)
    colunas_para_maiusculo = ["NOME DA EMPRESA", "TIPO DE MATERIAL", "TIPO DE SERVIÇO"]
    for col in colunas_para_maiusculo:
        if col in df_novo.columns:
            df_novo[col] = df_novo[col].astype(str).str.upper().str.strip()

    # 2. Carregar o arquivo original com openpyxl para manter os estilos
    arquivo_upload.seek(0)
    book = load_workbook(arquivo_upload)
    
    # Verifica se a aba existe, se não, cria uma baseada na primeira (ou cria nova)
    if nome_aba not in book.sheetnames:
        book.create_sheet(nome_aba)
    
    sheet = book[nome_aba]
    
    # 3. Descobrir a próxima linha vazia
    # Se a célula A1 estiver vazia, começamos na 1, senão na próxima disponível
    proxima_linha = sheet.max_row + 1
    if sheet.max_row == 1 and sheet['A1'].value is None:
        proxima_linha = 1

    # 4. Escrever os dados linha por linha (Preserva o estilo do arquivo)
    # Se for a primeira vez escrevendo (planilha vazia), escreve o cabeçalho
    if proxima_linha == 1:
        for c_idx, col_name in enumerate(df_novo.columns, 1):
            sheet.cell(row=1, column=c_idx, value=col_name)
        proxima_linha = 2

    # Escrever os valores do novo fornecedor
    for r_idx, row in df_novo.iterrows():
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=proxima_linha + r_idx, column=c_idx, value=value)

    # 5. Ajuste de largura de colunas (Sua lógica preferida)
    for i, col in enumerate(df_novo.columns, 1):
        # Aqui ele tenta medir o tamanho baseado no que já existe na coluna
        column_letter = get_column_letter(i)
        current_width = sheet.column_dimensions[column_letter].width
        new_width = max(len(str(df_novo.iloc[0, i-1])), 15) # Mínimo de 15
        if new_width > current_width:
            sheet.column_dimensions[column_letter].width = new_width

    # 6. Salvar para BytesIO
    output = BytesIO()
    book.save(output)
    return output.getvalue()